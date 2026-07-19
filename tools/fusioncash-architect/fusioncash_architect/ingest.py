"""Ingestion of Oracle Fusion Cash Management configuration extracts.

Reads the five CM configuration .xlsx exports (Parse Rules, Matching Rules,
Tolerance Rules, Recon Rulesets, Transaction Creation Rules) into normalized
Python dicts.

Design decisions (deliberate, and stricter than the PRD's pandas suggestion):
- Every cell is read as a STRING via openpyxl. GL account strings
  ("01-1100098-000000-100530-000...") and BAI transaction codes ("0142",
  "531") must never be float-coerced, which is exactly what pandas would do on
  ingest. Identifier fidelity is non-negotiable — see the reconciliation
  engines' no-pandas rule.
- Extracts carry a preamble (report title/filters) above the header row, a
  leading blank column, and abbreviated headers. The header row is LOCATED by
  scanning for known anchor tokens, never assumed at row 0.
- Personally identifying fields (the "last updated by" user id / consultant
  email) are dropped on ingest; the analysis never needs them and they should
  not travel into a generated report.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


class IngestError(Exception):
    """Raised when a file cannot be recognized or a required column is absent."""


def _norm_header(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().upper().split())


# Normalized header maps, keyed by the abbreviated header text Oracle exports.
# Anchor tokens (used to detect the file type + locate the header row) are the
# keys marked in DETECTORS below. "last updated by" columns are intentionally omitted.
_PARSE_MAP = {
    "PRS RLE SET": "rule_set", "DSCRPT": "description", "RLE SET ENBLD": "set_enabled",
    "SQNCENUM": "sequence", "ENBLD": "enabled", "TRX CDE": "txn_code",
    "TRX TPE": "txn_type", "SRCE FLD": "source_field", "TRGT FLD": "target_field",
    "PRSE RLE": "parse_rule", "OVRWRTE": "overwrite",
}
_MATCHING_MAP = {
    "MTCH RLE NME": "name", "DSCRPT": "description", "ENBLD": "enabled",
    "TRX SRCE": "source", "MTCH TPE": "match_type", "AMNT MTCH": "amount_match",
    "DTE MTCH": "date_match", "REF ID MTCH": "ref_id_match", "TPE MTCH": "type_match",
    "STMT GRPBY": "stmt_groupby", "TRX GRPBY": "txn_groupby", "ADV CRTRA": "advanced_criteria",
}
_TOLERANCE_MAP = {
    "TOL RLE NME": "name", "DSCRPT": "description", "DTE ENBLD": "date_enabled",
    "DAYS BFR": "days_before", "DAYS AFTR": "days_after", "AMNT ENBLD": "amount_enabled",
    "AMNT BELOW": "amount_below", "AMNT ABOVE": "amount_above", "PRCNT ENBLD": "pct_enabled",
    "PRCNT BELOW": "pct_below", "PRCNT ABOVE": "pct_above",
}
_RECON_MAP = {
    "RLST NME": "ruleset", "RLST DSCRPT": "description", "SQNCE NUM": "sequence",
    "MTCH RLE NME": "matching_rule", "TOL RLE NME": "tolerance_rule",
}
_TCR_MAP = {
    "BNK ACCNTNME": "bank_account", "ENBLD": "enabled", "SQC NUM": "sequence",
    "RLE NME": "name", "DSCRPT": "description", "TRX CDE": "txn_code",
    "TRX TPE": "txn_type", "SRCH FLD": "search_field", "SRCH STRNG": "search_string",
    "CASH": "cash_gl", "OFFSET": "offset_gl", "ACCNT FLG": "account_flag",
}

# kind -> (header map, anchor tokens that must all be present on the header row)
DETECTORS = {
    "parse": (_PARSE_MAP, {"PRSE RLE", "TRGT FLD"}),
    "matching": (_MATCHING_MAP, {"MTCH RLE NME", "ADV CRTRA"}),
    "tolerance": (_TOLERANCE_MAP, {"TOL RLE NME", "DAYS BFR"}),
    "recon": (_RECON_MAP, {"RLST NME", "SQNCE NUM"}),
    "tcr": (_TCR_MAP, {"BNK ACCNTNME", "CASH"}),
}


@dataclass
class Extract:
    kind: str
    source_file: str
    records: list[dict] = field(default_factory=list)


def _detect_kind(header_tokens: set[str]) -> str | None:
    for kind, (_, anchors) in DETECTORS.items():
        if anchors <= header_tokens:
            return kind
    return None


def _find_header(ws, scan_rows: int = 15):
    """Return (row_index, kind, {col_index: field_name}) or raise IngestError."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True)):
        tokens = {_norm_header(c) for c in row if c not in (None, "")}
        kind = _detect_kind(tokens)
        if kind:
            header_map = DETECTORS[kind][0]
            col_map = {}
            for ci, cell in enumerate(row):
                norm = _norm_header(cell)
                if norm in header_map:
                    col_map[ci] = header_map[norm]
            return i, kind, col_map
    raise IngestError(
        f"no recognizable CM configuration header found in the first {scan_rows} rows"
    )


def load_extract(path: str | Path) -> Extract:
    """Load one CM configuration .xlsx into a normalized Extract. Fail loud on
    an unrecognizable file rather than emitting phantom rows."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        header_i, kind, col_map = _find_header(ws)
        records = []
        for row in ws.iter_rows(min_row=header_i + 2, values_only=True):
            rec = {}
            for ci, name in col_map.items():
                val = row[ci] if ci < len(row) else None
                rec[name] = "" if val is None else str(val).strip()
            # Skip rows with no meaningful content (blank separator rows).
            if any(v for v in rec.values()):
                records.append(rec)
    finally:
        wb.close()
    return Extract(kind=kind, source_file=path.name, records=records)


def load_all(paths) -> dict[str, Extract]:
    """Load several extracts and key them by detected kind. A duplicate kind
    fails loud — two files claiming the same role is an upload error."""
    out: dict[str, Extract] = {}
    for p in paths:
        ex = load_extract(p)
        if ex.kind in out:
            raise IngestError(
                f"two files both parsed as '{ex.kind}': "
                f"{out[ex.kind].source_file} and {ex.source_file}"
            )
        out[ex.kind] = ex
    return out
