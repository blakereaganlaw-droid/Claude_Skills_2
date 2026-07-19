import openpyxl
import pytest

from fusioncash_architect.ingest import load_extract, load_all, IngestError


def _write_matching(path):
    """Mimic the real export layout: a preamble, a leading blank column, and
    abbreviated headers on a row that is not row 1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Report: Matching Rules"])          # preamble
    ws.append([])                                   # blank
    ws.append(["Filter: BU = UT"])                 # preamble
    ws.append([])
    ws.append([])
    ws.append([])
    # header row (index 6), leading blank column
    ws.append([None, "MTCH RLE NME", "DSCRPT", "MTCH TPE", "ADV CRTRA"])
    ws.append([None, "R1", "one to one", "1:1", "s.X LIKE '%A%'"])
    ws.append([None, "", "", "", ""])              # blank separator
    ws.append([None, "R2", "receivables", "1:M", "s.Y LIKE '%B%'"])
    wb.save(path)


def _write_tcr(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Report: TCR"])
    for _ in range(5):
        ws.append([])
    ws.append([None, "BNK ACCNTNME", "TRX TPE", "RLE NME", "CASH", "OFFSET"])
    ws.append([None, "BOA - Master", "BKF", "Bank Fee", "01-1100098-000", "01-6500000-000"])
    wb.save(path)


def test_detects_and_parses_matching(tmp_path):
    p = tmp_path / "m.xlsx"
    _write_matching(p)
    ex = load_extract(p)
    assert ex.kind == "matching"
    assert len(ex.records) == 2                      # blank separator dropped
    assert ex.records[0]["name"] == "R1"
    assert ex.records[0]["match_type"] == "1:1"
    assert ex.records[1]["match_type"] == "1:M"


def test_reference_strings_not_float_coerced(tmp_path):
    p = tmp_path / "t.xlsx"
    _write_tcr(p)
    ex = load_extract(p)
    assert ex.records[0]["cash_gl"] == "01-1100098-000"  # exact string preserved


def test_load_all_keys_by_kind(tmp_path):
    m, t = tmp_path / "m.xlsx", tmp_path / "t.xlsx"
    _write_matching(m)
    _write_tcr(t)
    out = load_all([m, t])
    assert set(out) == {"matching", "tcr"}


def test_unrecognized_file_fails_loud(tmp_path):
    p = tmp_path / "junk.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["totally", "unrelated", "columns"])
    wb.save(p)
    with pytest.raises(IngestError):
        load_extract(p)


def test_duplicate_kind_fails_loud(tmp_path):
    a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    _write_matching(a)
    _write_matching(b)
    with pytest.raises(IngestError):
        load_all([a, b])
